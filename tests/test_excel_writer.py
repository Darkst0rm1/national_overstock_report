from datetime import date

import pandas as pd

from src import config
from src.allocation import aggregate_allocations, build_allocation
from src.eligibility import eligible_material_universe, filter_relevant_sales_orders
from src.excel_writer import build_workbook
from src.old_report import build_food_service_report, build_old_report
from src.overstock import build_overstock
from src.price_list import build_price_list

REPORT_DATE = date(2026, 9, 1)


def _tiny_sales_order_df():
    row = {c: None for c in config.OVERSTOCK_COLUMNS}
    row.update(
        {
            "Sales Order": "3000001",
            "Sales Order Item": "10",
            "Material": "10000001",
            "Material Description": "TEST ITEM",
            "Confirmed Quantity (CS)": 5,
            "Plant": "2910",
        }
    )
    return pd.DataFrame([row])


def _tiny_pricing_df():
    return pd.DataFrame(
        [{"Material": "10000001", "BDM": "TEST BDM", "Pack": 1, "Size": "1 EA", "Sold by Wt": 0}]
    )


def _tiny_materials_df():
    return pd.DataFrame(
        [
            {
                "Material": "10000001",
                "Material Description": "TEST ITEM",
                "Plant": "2910",
                "Plant Name": "TOL Mississauga",
                "Batch": "0001",
                "Storage Location": "1100",
                "Shelf Life Expiration Date": pd.Timestamp("2026-09-01"),
                "Unrestricted Stock": 10,
            }
        ]
    )


def _build(sales_order_df=None, materials_df=None, pricing_df=None, report_date=REPORT_DATE):
    sales_order_df = _tiny_sales_order_df() if sales_order_df is None else sales_order_df
    materials_df = _tiny_materials_df() if materials_df is None else materials_df
    pricing_df = _tiny_pricing_df() if pricing_df is None else pricing_df

    eligible_materials = eligible_material_universe(materials_df, pricing_df)
    relevant_sales_order_df = filter_relevant_sales_orders(sales_order_df, eligible_materials)

    allocation = build_allocation(relevant_sales_order_df)
    allocation_totals = aggregate_allocations(relevant_sales_order_df)
    overstock_df = build_overstock(sales_order_df)
    price_list_df = build_price_list(pricing_df)
    old_report_df = build_old_report(materials_df, pricing_df, allocation_totals, report_date)
    food_service_df = build_food_service_report(materials_df, pricing_df, allocation_totals)
    return build_workbook(overstock_df, price_list_df, old_report_df, allocation, food_service_df)


def test_sheet_order_and_names():
    wb = _build()
    assert wb.sheetnames == ["Allocation", "Overstock", "Price List", "Old report", "Foodservice"]


def test_pivot_set_to_refresh_live_and_sorted_ascending():
    wb = _build()
    pt = wb["Allocation"]._pivots[0]
    assert pt.cache.refreshOnLoad is True
    assert pt.pivotFields[config.OVERSTOCK_COLUMNS.index("Material")].sortType == "ascending"
    assert pt.pivotFields[config.OVERSTOCK_COLUMNS.index("Plant")].sortType == "ascending"


def test_allocation_sheet_pivot_location_and_cells_are_left_for_excel_to_refresh():
    # Verified against real Excel (COM): pre-editing the pivot's own
    # location.ref/header/data cells before its refreshOnLoad-triggered
    # refresh runs makes that refresh collapse to a single #SPILL! cell
    # instead of resizing cleanly. Excel's native refresh already
    # resizes/repopulates the pivot correctly on its own, so
    # _write_allocation_sheet must leave the pivot's own output area alone
    # and only configure the cache (record count, sort order, refresh flag).
    wb = _build()
    ws = wb["Allocation"]
    pt = ws._pivots[0]
    assert pt.location.ref == "A3:G117"  # untouched template value, not shrunk to this run's size


def test_old_report_allocated_is_a_plain_value_not_a_pivot_formula():
    # Verified against real Excel (COM), including a genuine interactive
    # Ctrl+Alt+F9 full recalculation: ANY formula referencing the Allocation
    # pivot's output cells -- plain cell reference, INDEX/MATCH, or even
    # GETPIVOTDATA -- gets stuck at a stale/error-caught value from before
    # the pivot's refreshOnLoad refresh completes, and nothing short of
    # literally re-entering the formula fixes it. So Allocated is written
    # as a plain number, FIFO-computed the same way FIFO consumption always
    # is (src/fifo_allocation.py), which has no such failure mode.
    wb = _build()
    ws = wb["Old report"]
    cell = ws.cell(2, 14)  # A=BDM,B=Material,...,N=Allocated
    assert cell.data_type == "n"  # numeric constant, not a formula ("f")
    assert cell.value == 5  # tiny_sales_order_df: Material 10000001 @ 2910 confirmed qty 5


def test_old_report_allocation_applies_city_pairing_for_3pl_plants():
    # 3PL plants (2925/2935) don't take direct orders -- their Allocated
    # figure must come from their city's TOL plant pool instead.
    sales_order_df = pd.DataFrame(
        [
            {**{c: None for c in config.OVERSTOCK_COLUMNS}, "Sales Order": "1", "Sales Order Item": "10",
             "Material": "10000001", "Material Description": "X", "Confirmed Quantity (CS)": 9, "Plant": "2920"},
        ]
    )
    materials_df = pd.DataFrame(
        [
            {
                "Material": "10000001", "Material Description": "X",
                "Plant": "2925", "Plant Name": "Lineage Calgary", "Batch": "B1", "Storage Location": "1100",
                "Shelf Life Expiration Date": pd.Timestamp("2026-09-01"), "Unrestricted Stock": 10,
            }
        ]
    )
    wb = _build(sales_order_df=sales_order_df, materials_df=materials_df)

    ws = wb["Old report"]
    assert ws.cell(2, 14).value == 9  # 2925's own row picks up 2920's confirmed quantity


def test_old_report_allocation_is_consumed_once_across_batches_not_broadcast():
    # Reproduces the reported bug case: material with two batches at the
    # same plant must split one allocation total between them, FIFO by
    # expiration date, instead of applying the full total to each.
    sales_order_df = pd.DataFrame(
        [
            {**{c: None for c in config.OVERSTOCK_COLUMNS}, "Sales Order": "1", "Sales Order Item": "10",
             "Material": "10026066", "Material Description": "X", "Confirmed Quantity (CS)": 431, "Plant": "2910"},
        ]
    )
    materials_df = pd.DataFrame(
        [
            {
                "Material": "10026066", "Material Description": "X", "Plant": "2910",
                "Plant Name": "TOL Mississauga", "Batch": "EARLY", "Storage Location": "1100",
                "Shelf Life Expiration Date": pd.Timestamp("2026-09-02"), "Unrestricted Stock": 91,
            },
            {
                "Material": "10026066", "Material Description": "X", "Plant": "2910",
                "Plant Name": "TOL Mississauga", "Batch": "LATE", "Storage Location": "1100",
                "Shelf Life Expiration Date": pd.Timestamp("2026-09-10"), "Unrestricted Stock": 340,
            },
        ]
    )
    pricing_df = pd.DataFrame(
        [{"Material": "10026066", "BDM": "TEST BDM", "Pack": 1, "Size": "1 EA", "Sold by Wt": 0}]
    )
    wb = _build(sales_order_df=sales_order_df, materials_df=materials_df, pricing_df=pricing_df)

    ws = wb["Old report"]
    allocated = [ws.cell(r, 14).value for r in (2, 3)]
    stock = [ws.cell(r, 12).value for r in (2, 3)]
    assert sum(allocated) == 431  # consumed once, not double-counted
    assert all(a <= s for a, s in zip(allocated, stock))  # never over-allocates a batch
    assert ws.cell(2, 15).value == "=L2-N2"  # Available formula, per-row, unaffected


def test_old_report_has_dsd_and_deal_price_formulas():
    wb = _build()
    ws = wb["Old report"]
    assert ws.cell(2, 16).value == "=_xlfn.XLOOKUP(B:B,'Price List'!C:C,'Price List'!H:H)"
    assert ws.cell(2, 17).value == "=P2*(1-0.75)"


def test_old_report_unique_key_uses_mapped_plant_not_a_constant():
    materials_df = pd.DataFrame(
        [
            {
                "Material": "10000001", "Material Description": "X", "Plant": "2930",
                "Plant Name": "TOL Surrey", "Batch": "B1", "Storage Location": "1100",
                "Shelf Life Expiration Date": pd.Timestamp("2026-09-01"), "Unrestricted Stock": 10,
            }
        ]
    )
    pricing_df = pd.DataFrame(
        [{"Material": "10000001", "BDM": "TEST BDM", "Pack": 1, "Size": "1 EA", "Sold by Wt": 0}]
    )
    sales_order_df = pd.DataFrame(columns=config.OVERSTOCK_COLUMNS)
    wb = _build(sales_order_df=sales_order_df, materials_df=materials_df, pricing_df=pricing_df)

    ws = wb["Old report"]
    assert ws.cell(2, 13).value == "10000001_2930"  # own plant, not a hardcoded "_2910"


def test_foodservice_sheet_populated_from_excluded_bdm_materials():
    materials_df = pd.DataFrame(
        [
            {
                "Material": "10026066", "Material Description": "RANA ITEM", "Plant": "2910",
                "Plant Name": "TOL Mississauga", "Batch": "FS1", "Storage Location": "1100",
                "Shelf Life Expiration Date": pd.Timestamp("2020-01-01"), "Unrestricted Stock": 40,
            }
        ]
    )
    pricing_df = pd.DataFrame(
        [{"Material": "10026066", "BDM": "MIGUEL GUTIERREZ", "Pack": 1, "Size": "1 EA", "Sold by Wt": 0}]
    )
    sales_order_df = pd.DataFrame(columns=config.OVERSTOCK_COLUMNS)
    wb = _build(sales_order_df=sales_order_df, materials_df=materials_df, pricing_df=pricing_df)

    ws = wb["Foodservice"]
    assert ws.cell(2, 2).value == "10026066"
    assert ws.cell(2, 9).value == "FS1"
    # Excluded from the eligible-materials universe, so no Old report row.
    assert wb["Old report"].max_row == 1  # header only


def test_workbook_reopens_without_error():
    import io

    wb = _build()
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from openpyxl import load_workbook

    reopened = load_workbook(buf)
    assert reopened.sheetnames == ["Allocation", "Overstock", "Price List", "Old report", "Foodservice"]
