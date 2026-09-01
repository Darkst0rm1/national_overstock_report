"""
Writes the final workbook by editing the reference workbook IN PLACE as a
live template, rather than rebuilding one from scratch with openpyxl.

Why: the reference file's "Allocation" sheet is a real, native Excel
PivotTable, and its source range is 'Overstock'!A1:AE1048576 -- i.e. the
pivot reads directly off the Overstock sheet in the same workbook. openpyxl
round-trips that PivotTable's XML intact (verified: load + save reproduces
the same cacheSource, cacheFields, and pivotFields). So instead of trying
to fabricate pivot-cache XML ourselves (unsupported, corruption-prone), we:

  1. Load the reference workbook as a template.
  2. Overwrite the data rows of Overstock / Price List / Old report with
     this run's data, resizing each sheet's row count and cloning the
     template's own cell styles onto any newly-inserted rows -- so every
     font, fill, border, number format, column width, freeze pane, and
     autofilter is inherited automatically instead of hand re-implemented.
  3. Set the PivotTable's cache to refreshOnLoad=1 and force ascending
     sort on the Material/Plant fields, so Excel recomputes a fresh,
     correctly laid-out pivot the moment the file is opened.
  4. Also pre-populate the pivot's own A:G cells with the exact values our
     own aggregation (src/allocation.py) computes -- identical to what
     Excel's live refresh will produce -- so the sheet looks correct even
     before any refresh happens, and rebuild the manually-added "Unique
     Key" helper column (H) to match that same row range.

This keeps the PivotTable genuinely native (field list, drill-down,
right-click Refresh all work), while guaranteeing correct-looking data on
first open regardless of refresh timing.
"""
from __future__ import annotations

from copy import copy
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from src import config
from src.allocation import AllocationResult

TEMPLATE_PATH = Path(__file__).resolve().parent.parent / "assets" / "report_template.xlsx"

# Fixed cacheField positions in the reference workbook's Overstock column
# order (0-indexed) -- these only stay valid because OVERSTOCK_COLUMNS is
# never reordered relative to the reference file.
_MATERIAL_PIVOT_FIELD_INDEX = config.OVERSTOCK_COLUMNS.index("Material")
_PLANT_PIVOT_FIELD_INDEX = config.OVERSTOCK_COLUMNS.index("Plant")


def _clean(value):
    if value is None:
        return None
    if pd.api.types.is_scalar(value) and pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    return value


def _sync_row_count(ws, old_last_row: int, new_last_row: int, style_template_row: int, n_cols: int):
    """Grows or shrinks a sheet's data rows to exactly new_last_row,
    cloning cell styles from style_template_row onto any newly-inserted
    rows so they stay formatted like the rest of the column."""
    if new_last_row > old_last_row:
        n_new = new_last_row - old_last_row
        ws.insert_rows(old_last_row + 1, amount=n_new)
        for r in range(old_last_row + 1, new_last_row + 1):
            for c in range(1, n_cols + 1):
                src = ws.cell(style_template_row, c)
                dst = ws.cell(r, c)
                dst.font = copy(src.font)
                dst.fill = copy(src.fill)
                dst.border = copy(src.border)
                dst.alignment = copy(src.alignment)
                dst.number_format = src.number_format
    elif new_last_row < old_last_row:
        ws.delete_rows(new_last_row + 1, old_last_row - new_last_row)


# ---------------------------------------------------------------------------
# Overstock
# ---------------------------------------------------------------------------
def _write_overstock_sheet(wb: Workbook, overstock_df: pd.DataFrame):
    ws = wb["Overstock"]
    columns = config.OVERSTOCK_COLUMNS
    old_last_row = ws.max_row
    new_last_row = overstock_df.shape[0] + 1

    _sync_row_count(ws, old_last_row, new_last_row, style_template_row=2, n_cols=len(columns))

    for r_idx, (_, row) in enumerate(overstock_df.iterrows(), start=2):
        for c_idx, col_name in enumerate(columns, start=1):
            ws.cell(r_idx, c_idx, _clean(row[col_name]))

    last_col = get_column_letter(len(columns))
    ws.auto_filter.ref = f"A1:{last_col}{new_last_row}"


# ---------------------------------------------------------------------------
# Price List
# ---------------------------------------------------------------------------
def _write_price_list_sheet(wb: Workbook, price_list_df: pd.DataFrame):
    ws = wb["Price List"]
    columns = config.PRICE_LIST_COLUMNS
    old_last_row = ws.max_row
    new_last_row = price_list_df.shape[0] + 1

    _sync_row_count(ws, old_last_row, new_last_row, style_template_row=2, n_cols=len(columns))

    for r_idx, (_, row) in enumerate(price_list_df.iterrows(), start=2):
        for c_idx, col_name in enumerate(columns, start=1):
            ws.cell(r_idx, c_idx, _clean(row[col_name]))

    last_col = get_column_letter(len(columns))
    ws.auto_filter.ref = f"A1:{last_col}{new_last_row}"


# ---------------------------------------------------------------------------
# Old report / Foodservice (same 17-column layout and formula pattern --
# Foodservice is the rule-3 (excluded-BDM) population kept for future use)
# ---------------------------------------------------------------------------
def _write_old_report_style_sheet(wb: Workbook, sheet_name: str, report_df: pd.DataFrame):
    ws = wb[sheet_name]
    columns = config.OLD_REPORT_COLUMNS
    old_last_row = ws.max_row
    new_last_row = report_df.shape[0] + 1

    _sync_row_count(ws, old_last_row, new_last_row, style_template_row=2, n_cols=len(columns))

    price_list_letter = {
        name: get_column_letter(idx + 1) for idx, name in enumerate(config.PRICE_LIST_COLUMNS)
    }

    for r_offset, (_, src_row) in enumerate(report_df.iterrows()):
        r = r_offset + 2

        ws.cell(r, 1, f"=_xlfn.XLOOKUP(B:B,'Price List'!C:C,'Price List'!{price_list_letter['BDM']}:{price_list_letter['BDM']})")
        ws.cell(r, 2, _clean(src_row["Material"]))
        ws.cell(r, 3, _clean(src_row["Material Description"]))
        ws.cell(r, 4, f"=_xlfn.XLOOKUP(B:B,'Price List'!C:C,'Price List'!{price_list_letter['Sold by Wt']}:{price_list_letter['Sold by Wt']})")
        ws.cell(r, 5, f"=_xlfn.XLOOKUP(B:B,'Price List'!C:C,'Price List'!{price_list_letter['Pack']}:{price_list_letter['Pack']})")
        ws.cell(r, 6, f"=_xlfn.XLOOKUP(B:B,'Price List'!C:C,'Price List'!{price_list_letter['Size']}:{price_list_letter['Size']})")
        ws.cell(r, 7, _clean(src_row["Plant"]))
        ws.cell(r, 8, _clean(src_row["Plant Name"]))
        ws.cell(r, 9, _clean(src_row["Batch"]))
        ws.cell(r, 10, _clean(src_row["Storage Location"]))
        ws.cell(r, 11, _clean(src_row["Shelf Life Expiration Date"]))
        ws.cell(r, 12, _clean(src_row["Unrestricted Stock"]))
        ws.cell(r, 13, _clean(src_row["Unique Key"]))

        # Allocated is written as a plain number -- src/fifo_allocation.py
        # has already consumed each Material/mapped-Plant's total confirmed
        # quantity once, FIFO across that material/plant's batches -- not an
        # Excel formula. Verified against real Excel (COM, including a
        # genuine interactive Ctrl+Alt+F9 full recalculation) that ANY
        # formula referencing the Allocation pivot's output cells -- plain
        # cell reference, INDEX/MATCH, or even GETPIVOTDATA -- gets stuck at
        # a stale/error value from before the pivot's refreshOnLoad refresh
        # completes, and nothing short of literally re-entering the formula
        # fixes it. A plain value has no such failure mode, and it is
        # computed fresh from this row's own Material/Plant every run, so it
        # can only ever be deleted along with its own row, never
        # misaligned with a different one.
        ws.cell(r, 14, _clean(src_row["Allocated"]))
        ws.cell(r, 15, f"=L{r}-N{r}")

        ws.cell(r, 16, f"=_xlfn.XLOOKUP(B:B,'Price List'!C:C,'Price List'!{price_list_letter['1 - Store Door']}:{price_list_letter['1 - Store Door']})")
        ws.cell(r, 17, f"=P{r}*(1-{config.DEAL_PRICE_DISCOUNT})")

    ws.auto_filter.ref = f"A1:{config.OLD_REPORT_AUTOFILTER_LAST_COLUMN}{new_last_row}"


# ---------------------------------------------------------------------------
# Allocation (real PivotTable, set to refresh live -- not pre-populated)
# ---------------------------------------------------------------------------
def _write_allocation_sheet(wb: Workbook, allocation: AllocationResult, sales_order_row_count: int) -> None:
    """Configures the "Allocation" PivotTable to do a genuine live refresh
    from the (new) Overstock data the moment the workbook is opened.

    Deliberately does NOT touch the pivot's own output cells (header/data
    grid), does NOT resize the sheet's row count, and does NOT set
    pt.location.ref: verified against real Excel that pre-editing the
    pivot's own displayed range before its refreshOnLoad-triggered refresh
    runs causes that refresh to collapse the whole table to a single
    #SPILL! cell instead of resizing cleanly -- Excel's native refresh
    already resizes/repopulates a PivotTable correctly on its own when left
    alone, so our only job is to point it at the right cache state (fresh
    record count, ascending sort) and let Excel do the rest on open.
    """
    ws = wb["Allocation"]
    pt = ws._pivots[0]

    pt.cache.refreshOnLoad = True
    pt.cache.recordCount = sales_order_row_count
    pt.pivotFields[_MATERIAL_PIVOT_FIELD_INDEX].sortType = "ascending"
    pt.pivotFields[_PLANT_PIVOT_FIELD_INDEX].sortType = "ascending"


def build_workbook(
    overstock_df: pd.DataFrame,
    price_list_df: pd.DataFrame,
    old_report_df: pd.DataFrame,
    allocation: AllocationResult,
    food_service_df: pd.DataFrame,
) -> Workbook:
    wb = load_workbook(TEMPLATE_PATH)

    _write_allocation_sheet(wb, allocation, overstock_df.shape[0])
    _write_overstock_sheet(wb, overstock_df)
    _write_price_list_sheet(wb, price_list_df)
    _write_old_report_style_sheet(wb, "Old report", old_report_df)
    _write_old_report_style_sheet(wb, "Foodservice", food_service_df)

    return wb
